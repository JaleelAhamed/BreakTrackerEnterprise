"""
Break Tracker Enterprise
Report Generator Module

Version: 2.0.0

Generates professional Excel productivity reports using OpenPyXL.

Public API
----------
generate_session_report(employee, session, break_log, allowed_break_minutes) -> Path
    This is the stable entry point consumed by ``session.py``. Its
    signature and return type must not change.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app_paths import REPORTS_DIR
from logger import get_logger

logger = get_logger(__name__)


# ======================================================================
# Design tokens
# ======================================================================
#
# All colors, spacing, and geometry used by the report live here so that
# the visual design can be tuned in one place without touching the
# report-building logic below.


@dataclass(frozen=True)
class _Palette:
    """Enterprise color palette used throughout the workbook."""

    title_navy: str = "0F2D4D"      # main report title banner
    subtitle_blue: str = "2E5F8A"   # title sub-line banner
    section_blue: str = "1F4E78"    # section header banners
    table_header_blue: str = "17365D"  # break-details table header (darkest)
    label_bg: str = "F2F2F2"        # label cell background
    zebra_fill: str = "DCE6F1"      # alternating row tint
    border_gray: str = "BFBFBF"
    footer_gray: str = "595959"
    white: str = "FFFFFF"
    black: str = "000000"


@dataclass(frozen=True)
class _Layout:
    """Shared geometry: column spans, row heights, and spacing."""

    section_span: int = 6       # columns spanned by titles & section headers
    table_span: int = 5         # columns spanned by the break-details table

    title_row_height: int = 30
    subtitle_row_height: int = 20
    section_row_height: int = 22
    label_row_height: int = 19
    table_header_row_height: int = 21
    table_row_height: int = 19
    spacer_row_height: int = 8

    column_padding: int = 4     # extra width added by auto-fit
    min_column_width: int = 12
    max_column_width: int = 45

    default_column_widths: dict = field(
        default_factory=lambda: {1: 24, 2: 20, 3: 20, 4: 20, 5: 20, 6: 20}
    )


PALETTE = _Palette()
LAYOUT = _Layout()


class _StyleKit:
    """
    Builds and caches every Font / Fill / Border / Alignment the report
    needs, once, so call sites reference a shared style object by name
    instead of constructing (and duplicating) style objects inline.
    """

    def __init__(self, palette: _Palette) -> None:
        self.thin_border = Border(
            left=Side(style="thin", color=palette.border_gray),
            right=Side(style="thin", color=palette.border_gray),
            top=Side(style="thin", color=palette.border_gray),
            bottom=Side(style="thin", color=palette.border_gray),
        )

        self.title_fill = self._solid_fill(palette.title_navy)
        self.subtitle_fill = self._solid_fill(palette.subtitle_blue)
        self.section_fill = self._solid_fill(palette.section_blue)
        self.table_header_fill = self._solid_fill(palette.table_header_blue)
        self.label_fill = self._solid_fill(palette.label_bg)
        self.zebra_fill = self._solid_fill(palette.zebra_fill)

        self.title_font = Font(name="Calibri", color=palette.white, bold=True, size=20)
        self.subtitle_font = Font(name="Calibri", color=palette.white, bold=True, size=12, italic=True)
        self.section_font = Font(name="Calibri", color=palette.white, bold=True, size=12)
        self.table_header_font = Font(name="Calibri", color=palette.white, bold=True, size=11)
        self.label_font = Font(name="Calibri", color=palette.black, bold=True, size=11)
        self.value_font = Font(name="Calibri", color=palette.black, size=11)
        self.footer_font = Font(name="Calibri", color=palette.footer_gray, italic=True, size=9)

        self.center = Alignment(horizontal="center", vertical="center")
        self.left = Alignment(horizontal="left", vertical="center", indent=1)
        self.right = Alignment(horizontal="right", vertical="center", indent=1)

    @staticmethod
    def _solid_fill(color: str) -> PatternFill:
        return PatternFill(fill_type="solid", start_color=color, end_color=color)


# ======================================================================
# Report generator
# ======================================================================


class ExcelReportGenerator:
    """
    Builds a single-sheet, professionally formatted Break Tracker
    Enterprise productivity report.

    Usage is sequential: each ``add_*`` method appends one section to the
    worksheet and advances an internal row cursor, so sections always
    appear in the order they are called.
    """

    def __init__(self) -> None:
        self.workbook = Workbook()
        self.sheet: Worksheet = self.workbook.active
        self.sheet.title = "Employee Report"

        self.styles = _StyleKit(PALETTE)
        self.current_row = 1
        self._section_row_index = 0

        self._configure_worksheet()

    # ------------------------------------------------------------
    # Worksheet-level configuration
    # ------------------------------------------------------------

    def _configure_worksheet(self) -> None:
        """Applies print, layout, and display settings for the sheet."""
        ws = self.sheet

        ws.sheet_view.showGridLines = False

        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.6
        ws.page_margins.bottom = 0.6
        ws.page_margins.header = 0.3
        ws.page_margins.footer = 0.3

        ws.print_options.horizontalCentered = True

        # Keep the title visible while scrolling through a long report.
        ws.freeze_panes = "A4"

        for column_index, width in LAYOUT.default_column_widths.items():
            ws.column_dimensions[get_column_letter(column_index)].width = width

    # ------------------------------------------------------------
    # Low-level cursor / cell helpers
    # ------------------------------------------------------------

    def _advance(self, rows: int = 1) -> None:
        """Move the row cursor forward."""
        self.current_row += rows

    def _add_spacer(self) -> None:
        """Writes a short blank row to give sections visual breathing room."""
        self.sheet.row_dimensions[self.current_row].height = LAYOUT.spacer_row_height
        self._advance()

    def _write_banner(
        self,
        text: str,
        *,
        font: Font,
        fill: PatternFill,
        alignment: Alignment,
        row_height: int,
    ) -> None:
        """
        Writes a single full-width, merged, filled banner row (used for
        both the report title and section headers) and advances the
        cursor past it.
        """
        start_row = self.current_row

        self.sheet.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=LAYOUT.section_span,
        )

        for column in range(1, LAYOUT.section_span + 1):
            cell = self.sheet.cell(row=start_row, column=column)
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment

        self.sheet.cell(row=start_row, column=1).value = text
        self.sheet.row_dimensions[start_row].height = row_height

        self._advance()

    def _write_label_value(self, label: str, value: Any) -> None:
        """Writes one bordered, merged ``label | value`` row with zebra shading."""
        row = self.current_row
        value_start_col = 2
        value_end_col = LAYOUT.section_span

        self.sheet.merge_cells(
            start_row=row,
            start_column=value_start_col,
            end_row=row,
            end_column=value_end_col,
        )

        label_cell = self.sheet.cell(row=row, column=1)
        label_cell.value = label
        label_cell.font = self.styles.label_font
        label_cell.fill = self.styles.label_fill
        label_cell.alignment = self.styles.left
        label_cell.border = self.styles.thin_border

        zebra = (self._section_row_index % 2 == 1)

        for column in range(value_start_col, value_end_col + 1):
            cell = self.sheet.cell(row=row, column=column)
            if column == value_start_col:
                cell.value = value
            cell.font = self.styles.value_font
            cell.alignment = self.styles.left
            cell.border = self.styles.thin_border
            if zebra:
                cell.fill = self.styles.zebra_fill

        self.sheet.row_dimensions[row].height = LAYOUT.label_row_height
        self._section_row_index += 1
        self._advance()

    def _write_merged_note(self, text: str, span: int) -> None:
        """Writes a single centered, bordered note spanning ``span`` columns."""
        row = self.current_row

        self.sheet.merge_cells(
            start_row=row,
            start_column=1,
            end_row=row,
            end_column=span,
        )

        for column in range(1, span + 1):
            cell = self.sheet.cell(row=row, column=column)
            cell.alignment = self.styles.center
            cell.border = self.styles.thin_border
            cell.font = self.styles.value_font

        self.sheet.cell(row=row, column=1).value = text
        self.sheet.row_dimensions[row].height = LAYOUT.table_row_height

        self._advance(2)

    # ------------------------------------------------------------
    # Section header helpers
    # ------------------------------------------------------------

    def _create_title_row(self, text: str, *, subtitle: bool = False) -> None:
        if subtitle:
            self._write_banner(
                text,
                font=self.styles.subtitle_font,
                fill=self.styles.subtitle_fill,
                alignment=self.styles.center,
                row_height=LAYOUT.subtitle_row_height,
            )
        else:
            self._write_banner(
                text,
                font=self.styles.title_font,
                fill=self.styles.title_fill,
                alignment=self.styles.center,
                row_height=LAYOUT.title_row_height,
            )

    def _create_section_header(self, title: str) -> None:
        self._write_banner(
            title,
            font=self.styles.section_font,
            fill=self.styles.section_fill,
            alignment=self.styles.left,
            row_height=LAYOUT.section_row_height,
        )
        self._section_row_index = 0

    # ------------------------------------------------------------
    # Table helpers
    # ------------------------------------------------------------

    def _write_table_header(self, headers: Sequence[str]) -> None:
        """Writes a bold, filled header row and enables auto-filter on it."""
        header_row = self.current_row

        for column, header in enumerate(headers, start=1):
            cell = self.sheet.cell(row=header_row, column=column)
            cell.value = header
            cell.font = self.styles.table_header_font
            cell.fill = self.styles.table_header_fill
            cell.border = self.styles.thin_border
            cell.alignment = self.styles.center

        self.sheet.row_dimensions[header_row].height = LAYOUT.table_header_row_height

        last_column = get_column_letter(len(headers))
        self.sheet.auto_filter.ref = f"A{header_row}:{last_column}{header_row}"

        self._advance()

    def _write_table_row(self, values: Sequence[Any], *, zebra: bool) -> None:
        """Writes one bordered, centered data row, optionally shaded."""
        for column, value in enumerate(values, start=1):
            cell = self.sheet.cell(row=self.current_row, column=column)
            cell.value = value
            cell.border = self.styles.thin_border
            cell.alignment = self.styles.center
            cell.font = self.styles.value_font

            if zebra:
                cell.fill = self.styles.zebra_fill

        self.sheet.row_dimensions[self.current_row].height = LAYOUT.table_row_height

        self._advance()

    # ------------------------------------------------------------
    # Formatting utilities
    # ------------------------------------------------------------

    @staticmethod
    def format_timedelta(duration: timedelta) -> str:
        """Converts a ``timedelta`` into an ``HH:MM:SS`` string."""
        total_seconds = int(duration.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours:02}:{minutes:02}:{seconds:02}"

    def auto_fit_columns(self) -> None:
        """Resizes every column to fit its widest cell's content, within sane bounds."""
        for column_cells in self.sheet.columns:
            column_letter = get_column_letter(column_cells[0].column)

            longest = 0
            for cell in column_cells:
                if cell.value is None:
                    continue
                longest = max(longest, len(str(cell.value)))

            computed_width = longest + LAYOUT.column_padding
            current_width = self.sheet.column_dimensions[column_letter].width or 0
            width = max(computed_width, current_width, LAYOUT.min_column_width)

            self.sheet.column_dimensions[column_letter].width = min(
                width, LAYOUT.max_column_width
            )

    # ------------------------------------------------------------
    # Report sections
    # ------------------------------------------------------------

    def create_report_title(self) -> None:
        """Writes the two-line enterprise report title."""
        self._create_title_row("BREAK TRACKER ENTERPRISE")
        self._create_title_row("Employee Productivity Report", subtitle=True)
        self._add_spacer()

    def add_employee_information(self, employee: Any, report_date: datetime) -> None:
        """Writes the Employee Information section."""
        self._create_section_header("Employee Information")

        self._write_label_value("Employee Name", getattr(employee, "name", ""))
        self._write_label_value("Employee ID", getattr(employee, "employee_id", ""))
        self._write_label_value("Department", getattr(employee, "department", ""))
        self._write_label_value("Designation", getattr(employee, "designation", ""))
        self._write_label_value("Report Date", report_date.strftime("%d-%b-%Y"))

        self._add_spacer()

    def add_session_summary(
        self,
        login_time: datetime,
        logout_time: datetime,
        session_duration: timedelta,
        productive_time: timedelta,
    ) -> None:
        """Writes the Session Summary section."""
        self._create_section_header("Session Summary")

        self._write_label_value("Login Time", login_time.strftime("%H:%M:%S"))
        self._write_label_value("Logout Time", logout_time.strftime("%H:%M:%S"))
        self._write_label_value(
            "Total Session Duration", self.format_timedelta(session_duration)
        )
        self._write_label_value(
            "Productive Time", self.format_timedelta(productive_time)
        )

        self._add_spacer()

    def add_break_summary(
        self,
        total_break: timedelta,
        allowed_break_minutes: int,
        exceeded_minutes: int,
        exceeded: bool,
    ) -> None:
        """Writes the Break Summary section."""
        self._create_section_header("Break Summary")

        self._write_label_value(
            "Total Break Duration", self.format_timedelta(total_break)
        )
        self._write_label_value("Allowed Break (Minutes)", allowed_break_minutes)
        self._write_label_value("Break Exceeded", "Yes" if exceeded else "No")
        self._write_label_value("Exceeded Minutes", exceeded_minutes)

        self._add_spacer()

    def add_break_details(self, break_events: Sequence[Any]) -> None:
        """Writes the detailed, per-break history table."""
        self._create_section_header("Break Details")

        headers = ["No", "Break Start", "Break End", "Duration", "Reason"]
        self._write_table_header(headers)

        if not break_events:
            self._write_merged_note(
                "No breaks recorded during this session.", LAYOUT.table_span
            )
            return

        for index, event in enumerate(break_events, start=1):
            row_values = [
                index,
                event.start_time.strftime("%H:%M:%S"),
                event.end_time.strftime("%H:%M:%S"),
                self.format_timedelta(event.duration),
                event.reason if event.reason else "Not Specified",
            ]
            self._write_table_row(row_values, zebra=(index % 2 == 0))

        self._add_spacer()

    def add_statistics(
        self,
        session_duration: timedelta,
        productive_time: timedelta,
        total_break: timedelta,
        break_events: Sequence[Any],
    ) -> float:
        """
        Writes the Productivity Statistics section and returns the
        computed productivity percentage for reuse in the Remarks section.
        """
        self._create_section_header("Productivity Statistics")

        break_count = len(break_events)

        if break_count > 0:
            longest_break = max(event.duration for event in break_events)
            average_break = total_break / break_count
        else:
            longest_break = timedelta(0)
            average_break = timedelta(0)

        productivity = 0.0
        if session_duration.total_seconds() > 0:
            productivity = (
                productive_time.total_seconds() / session_duration.total_seconds()
            ) * 100

        self._write_label_value("Break Count", break_count)
        self._write_label_value(
            "Longest Break", self.format_timedelta(longest_break)
        )
        self._write_label_value(
            "Average Break", self.format_timedelta(average_break)
        )
        self._write_label_value("Productivity %", f"{productivity:.2f} %")

        self._add_spacer()

        return productivity

    def add_remarks(self, productivity: float, exceeded: bool) -> None:
        """Writes automatically generated remarks and an overall status."""
        self._create_section_header("Remarks")

        remark = (
            "Employee exceeded the configured break allowance."
            if exceeded
            else "Employee remained within the configured break allowance."
        )
        status = self._productivity_status(productivity)

        self._write_label_value("Remarks", remark)
        self._write_label_value("Overall Status", status)

        self._add_spacer()

    @staticmethod
    def _productivity_status(productivity: float) -> str:
        """Maps a productivity percentage to a qualitative status label."""
        if productivity >= 90:
            return "Excellent"
        if productivity >= 75:
            return "Good"
        if productivity >= 60:
            return "Average"
        return "Needs Improvement"

    def add_footer(self) -> None:
        """Writes a clean, understated report metadata footer."""
        self._add_spacer()
        self._create_section_header("Report Information")

        self._write_label_value("Generated By", "Break Tracker Enterprise")
        self._write_label_value("Version", "2.0.0")
        self._write_label_value(
            "Generated On", datetime.now().strftime("%d-%b-%Y %I:%M:%S %p")
        )

        # Thin closing rule so the report doesn't end abruptly.
        closing_row = self.current_row
        self.sheet.merge_cells(
            start_row=closing_row,
            start_column=1,
            end_row=closing_row,
            end_column=LAYOUT.section_span,
        )
        closing_cell = self.sheet.cell(row=closing_row, column=1)
        closing_cell.value = "This report was generated automatically by Break Tracker Enterprise."
        closing_cell.font = self.styles.footer_font
        closing_cell.alignment = self.styles.center
        self.sheet.row_dimensions[closing_row].height = LAYOUT.spacer_row_height + 6

        self._advance()

    # ------------------------------------------------------------
    # Persistence
    # ------------------------------------------------------------

    def save(self, output_path: Path) -> Path:
        """Auto-fits columns, finalizes print layout, and writes the workbook."""
        self.auto_fit_columns()

        last_row = max(self.current_row - 1, 1)
        last_column_letter = get_column_letter(LAYOUT.section_span)

        self.sheet.print_title_rows = "1:2"
        self.sheet.print_area = f"A1:{last_column_letter}{last_row}"

        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(output_path)

        return output_path


# ======================================================================
# Public API
# ======================================================================


def _resolve_report_filename(employee: Any, timestamp: str) -> str:
    """Builds the report filename, tolerating either employee attribute name."""
    employee_name = getattr(employee, "employee_name", None) or getattr(
        employee, "name", "employee"
    )
    return f"{employee_name}_{timestamp}.xlsx"


def generate_session_report(
    employee: Any,
    session: Any,
    break_log: Any,
    allowed_break_minutes: int,
) -> Path:
    """
    Generates the employee session report.

    This function is the public API used by ``session.py`` and must keep
    this exact signature and return type.
    """
    employee_name = getattr(employee, "name", "<unknown>")
    logger.info("Report generation started for employee: %s", employee_name)

    try:
        session_duration: timedelta = session.duration()
        total_break: timedelta = break_log.total_duration()

        productive_time = session_duration - total_break
        if productive_time.total_seconds() < 0:
            productive_time = timedelta(0)

        exceeded_minutes = max(
            0,
            int(total_break.total_seconds() // 60) - allowed_break_minutes,
        )
        exceeded = exceeded_minutes > 0

        generator = ExcelReportGenerator()

        generator.create_report_title()

        generator.add_employee_information(
            employee=employee,
            report_date=datetime.now(),
        )

        generator.add_session_summary(
            login_time=session.login_time,
            logout_time=session.logout_time,
            session_duration=session_duration,
            productive_time=productive_time,
        )

        generator.add_break_summary(
            total_break=total_break,
            allowed_break_minutes=allowed_break_minutes,
            exceeded_minutes=exceeded_minutes,
            exceeded=exceeded,
        )

        generator.add_break_details(break_log.breaks)

        productivity = generator.add_statistics(
            session_duration=session_duration,
            productive_time=productive_time,
            total_break=total_break,
            break_events=break_log.breaks,
        )

        generator.add_remarks(productivity=productivity, exceeded=exceeded)

        generator.add_footer()

        logger.info("Employee report created for: %s", employee_name)

        reports_folder = REPORTS_DIR
        reports_folder.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = _resolve_report_filename(employee, timestamp)
        report_path = reports_folder / filename

        saved_path = generator.save(report_path)
        logger.info("Excel file saved: %s", saved_path)
        logger.info("Report path: %s", saved_path)

        return saved_path

    except Exception:
        logger.exception("Report generation failed for employee: %s", employee_name)
        raise